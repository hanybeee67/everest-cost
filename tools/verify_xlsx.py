# -*- coding: utf-8 -*-
"""생성된 엑셀의 모든 수식을 실제로 계산해서
   ① 오류값(#DIV/0!, #N/A, #REF! …)이 하나도 없는지
   ② JS 계산 엔진 결과와 값이 일치하는지 검증한다."""
import sys, json, warnings
warnings.filterwarnings('ignore')
import formulas

XLSX = sys.argv[1] if len(sys.argv) > 1 else 'public/downloads/에베레스트_원가분석서.xlsx'
EXPECT = json.load(open(sys.argv[2])) if len(sys.argv) > 2 else None

import openpyxl, re
_wb = openpyxl.load_workbook(XLSX)
FORMULA_CELLS = set()
for _ws in _wb.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            if isinstance(_c.value, str) and _c.value.startswith('='):
                FORMULA_CELLS.add(("'%s'!%s" % (_ws.title, _c.coordinate)).upper())
print('▶ 수식 셀 %d개 확인 대상' % len(FORMULA_CELLS))
print('▶ 엑셀 로딩 및 전체 수식 계산 중 …')
import tempfile, os, glob
xl = formulas.ExcelModel().loads(XLSX).finish()
xl.calculate()
_tmp = tempfile.mkdtemp()
xl.write(dirpath=_tmp)
_calc = glob.glob(os.path.join(_tmp, '*.[xX][lL][sS][xX]'))[0]
_cw = openpyxl.load_workbook(_calc, data_only=True)

ERR_TOKENS = ('#DIV/0!', '#N/A', '#REF!', '#VALUE!', '#NAME?', '#NUM!', '#NULL!', '#CYCLE!')
errors, values = [], {}
for ws in _cw.worksheets:
    for row in ws.iter_rows():
        for c in row:
            ref = ("'%s'!%s" % (ws.title, c.coordinate)).upper()
            values[ref] = c.value
            if ref in FORMULA_CELLS:
                sv = str(c.value)
                if any(t in sv for t in ERR_TOKENS):
                    errors.append((ref, sv))

print(f'   계산된 셀 참조: {len(values):,}개')
if errors:
    print(f'❌ 오류 셀 {len(errors)}개')
    for k, s in errors[:40]:
        print('   ', k, '=', s)
    sys.exit(1)
print('✅ 오류값(#DIV/0!, #N/A, #REF! 등) 0건')

if EXPECT:
    bad = []
    for ref, exp in EXPECT.items():
        got = values.get(ref.upper())
        if got is None:
            bad.append((ref, exp, 'MISSING')); continue
        try:
            g = float(got)
        except Exception:
            bad.append((ref, exp, got)); continue
        if abs(g - float(exp)) > max(1.0, abs(float(exp)) * 0.002):
            bad.append((ref, exp, g))
    if bad:
        print(f'❌ JS 엔진과 값이 다른 셀 {len(bad)}개')
        for b in bad[:30]:
            print('   ', b)
        sys.exit(1)
    print(f'✅ JS 계산 엔진과 대조: {len(EXPECT)}개 셀 전부 일치')

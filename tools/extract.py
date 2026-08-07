import openpyxl, json, re, os
COST='/root/.claude/uploads/dee33a02-e363-56dc-bbe7-a13aeae40c03/284592eb-________________1.xlsx'
RECIPE='/root/.claude/uploads/dee33a02-e363-56dc-bbe7-a13aeae40c03/6835d42d-_______v3.xlsx'
out={}

# ---------- recipes ----------
wb=openpyxl.load_workbook(RECIPE)
recipes=[]
for ws in wb.worksheets:
    rows=[[c.value for c in r] for r in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=7)]
    def g(r,c):
        v=rows[r-1][c-1]
        return '' if v is None else str(v).strip()
    name=g(2,2).strip()
    serving=g(2,7)
    cat=g(3,2)
    time=g(4,2)
    # ingredients: find header row where D=='번호'
    hdr=None
    for i in range(1,len(rows)+1):
        if g(i,4)=='번호': hdr=i;break
    ings=[]
    i=hdr+1
    while i<=len(rows):
        no=g(i,4); nm=g(i,5); qty=g(i,6); note=g(i,7)
        if no=='' and nm=='':
            break
        if nm: ings.append({'name':nm,'qty':qty,'note':note})
        i+=1
    # steps
    steps=[];garnish=''
    mode=None
    for r in range(1,len(rows)+1):
        b=g(r,2)
        if '조리 방법' in b: mode='s';continue
        if '가니쉬' in b: mode='g';continue
        if 'Recipe Card' in b: mode=None;continue
        if mode=='s' and b: steps.append(re.sub(r'^\d+\.\s*','',b).strip())
        elif mode=='g' and b: garnish=(garnish+' '+b).strip()
    recipes.append({'name':name,'sheet':ws.title,'category_raw':cat,'time':time,'serving':serving,
                    'ingredients':ings,'steps':steps,'garnish':garnish})
out['recipes']=recipes

# ---------- ingredient DB ----------
wb2=openpyxl.load_workbook(COST)
ws=wb2['🥬 식재료 DB']
ings=[]
for r in range(7,126):
    nm=ws.cell(r,3).value
    if not nm: continue
    ings.append({'name':str(nm).strip(),
                 'price':ws.cell(r,4).value or 0,
                 'pack_qty':ws.cell(r,5).value or 0,
                 'unit':ws.cell(r,6).value or 'kg',
                 'weight_g':ws.cell(r,7).value or 0})
out['ingredients']=ings

# ---------- fixed cost ----------
ws=wb2['💰 고정비용 DB']
fc=[]
for r in range(9,22):
    nm=ws.cell(r,3).value
    if not nm: continue
    fc.append({'name':str(nm).strip(),'amount':ws.cell(r,4).value or 0,'note':str(ws.cell(r,5).value or '')})
out['fixed_costs']=fc
out['monthly_revenue']=ws.cell(25,4).value

# ---------- prep ----------
ws=wb2['🧪 프렙 원가표']
preps=[]
for r in range(8,19):
    nm=ws.cell(r,3).value
    if not nm: continue
    preps.append({'name':str(nm).strip(),'yield_g':ws.cell(r,5).value,'yield_rate':ws.cell(r,6).value,'items':[]})
# details
cur=None
for r in range(20,ws.max_row+1):
    b=str(ws.cell(r,2).value or '')
    if '🧪' in b:
        t=b.replace('🧪','').strip()
        cur=next((p for p in preps if p['name']==t),None)
        continue
    if cur is not None:
        nm=ws.cell(r,3).value; qty=ws.cell(r,4).value
        lbl=str(ws.cell(r,2).value or '')
        if lbl.startswith('합'): cur=None; continue
        if nm and isinstance(qty,(int,float)):
            cur['items'].append({'name':str(nm).strip(),'qty':qty})
out['preps']=preps

# ---------- prices from old cost sheets ----------
prices={}
for t in wb2.sheetnames:
    if t in ('📊 대시보드','🥬 식재료 DB','💰 고정비용 DB','🧪 프렙 원가표'): continue
    w=wb2[t]
    for r in range(1,w.max_row+1):
        c=str(w.cell(r,3).value or '')
        if '🍽' in c:
            nm=c.replace('🍽','').strip()
            p=w.cell(r,5).value
            if isinstance(p,(int,float)): prices.setdefault(nm,p)
out['prices']=prices
json.dump(out,open('data/_raw.json','w'),ensure_ascii=False,indent=1)
print('recipes',len(recipes),'ings',len(ings),'preps',len(preps),'prices',len(prices))
